"""
XLSX Builder for Comprehensive Quality Report.
Creates a 10-sheet Excel workbook with data tables, embedded matplotlib charts,
and AI-powered insights per sheet.
"""

import io
import os
import tempfile
import shutil
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional

import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR
from analyzer import ComprehensiveResult
from xlsx_theme import (
    XlsxStyles, CHART_COLORS, ROW_HEIGHTS, COL_WIDTHS,
    format_duration, format_pct, format_number,
    sla_level, rating_text, efficiency_level, setup_chart_style,
)
from xlsx_analyzer import (
    XlsxDetailAnalyzer, ActionPlanRow,
)
from xlsx_visualizer import (
    chart_exec_health_gauge, chart_exec_process_radar, chart_exec_sparklines,
    chart_inc_monthly_trend, chart_inc_priority_pie, chart_inc_category_top10,
    chart_inc_mttr_boxplot, chart_inc_p1p2_trend,
    chart_sla_gauge_response, chart_sla_gauge_resolution,
    chart_sla_monthly_trend, chart_sla_violation_by_priority,
    chart_sla_violation_heatmap,
    chart_chg_type_pie, chart_chg_success_trend, chart_chg_category_bar,
    chart_chg_incident_scatter, chart_chg_planning_accuracy,
    chart_req_type_pie, chart_req_csat_bar, chart_req_monthly_trend,
    chart_req_fulfillment_boxplot, chart_req_dept_heatmap,
    chart_prb_status_funnel, chart_prb_rootcause_pie, chart_prb_monthly_bar,
    chart_prb_impact_bubble, chart_prb_rca_trend,
    chart_cross_sankey, chart_cross_radar, chart_cross_timeline,
    chart_cross_heatmap,
    chart_pers_workload_bar, chart_pers_load_boxplot, chart_pers_skill_heatmap,
    chart_pers_efficiency_scatter, chart_pers_top10_radar,
    chart_time_four_process_trend, chart_time_dow_bar, chart_time_hour_heatmap,
    chart_time_quarterly, chart_time_forecast,
    chart_action_priority_pie, chart_action_process_bar,
)

logger = logging.getLogger(__name__)

# ── Sheet names ──────────────────────────────────────────────────────────────

SHEET_NAMES = {
    "zh": [
        "执行摘要", "事件分析", "事件SLA分析", "变更分析", "请求分析",
        "问题分析", "跨流程关联分析", "人员与效率分析", "时间维度分析", "行动计划",
    ],
    "en": [
        "Executive Summary", "Incident Analysis", "Incident SLA",
        "Change Analysis", "Request Analysis", "Problem Analysis",
        "Cross-Process", "Personnel & Efficiency", "Time Analysis",
        "Action Plan",
    ],
}


class XlsxBuilder:
    """Assembles a 10-sheet Excel workbook from analysis results."""

    def __init__(
        self,
        result: ComprehensiveResult,
        incidents_df,
        changes_df,
        requests_df,
        problems_df,
        sla_map: Dict,
        insights: Dict[str, str],
        language: str = "en",
    ):
        self.result = result
        self.incidents_df = incidents_df
        self.changes_df = changes_df
        self.requests_df = requests_df
        self.problems_df = problems_df
        self.sla_map = sla_map or {}
        self.insights = insights or {}
        self.language = language

        self.styles = XlsxStyles(language)
        self.detail = XlsxDetailAnalyzer(
            incidents_df, changes_df, requests_df, problems_df,
            sla_map, result, language,
        )
        self.wb: Optional[openpyxl.Workbook] = None
        self._tmp_dir = tempfile.mkdtemp(prefix="xlsx_charts_")
        self._chart_counter = 0

        setup_chart_style(language)

    # ─── Utility methods ─────────────────────────────────────────────────

    def _write_title(self, ws, title: str, subtitle: str = "", row: int = 1) -> int:
        """Write sheet title and optional subtitle. Returns next available row."""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = self.styles.font_h1
        cell.alignment = self.styles.align_h1
        ws.row_dimensions[row].height = ROW_HEIGHTS["h1"]
        row += 1

        if not subtitle:
            subtitle = f"{self.result.start_date} — {self.result.end_date}"
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=1, value=subtitle)
        cell.font = self.styles.font_h2
        cell.alignment = self.styles.align_left
        ws.row_dimensions[row].height = ROW_HEIGHTS["h2"]
        row += 1

        # spacer
        ws.row_dimensions[row].height = ROW_HEIGHTS["spacer"]
        return row + 1

    def _write_section(self, ws, title: str, row: int) -> int:
        """Write a section heading. Returns next row."""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = self.styles.font_h3
        cell.alignment = self.styles.align_left
        ws.row_dimensions[row].height = ROW_HEIGHTS["h3"]
        return row + 1

    def _write_table(self, ws, headers: List[str], data: List[List], row: int,
                     col_widths: Optional[List[float]] = None) -> int:
        """Write a table with header and data rows. Returns next row after table."""
        # Header row
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = self.styles.font_th
            cell.fill = self.styles.fill_header
            cell.alignment = self.styles.align_center
            cell.border = self.styles.border_header
        ws.row_dimensions[row].height = ROW_HEIGHTS["th"]
        row += 1

        # Data rows with zebra striping
        for ri, data_row in enumerate(data):
            fill = self.styles.fill_zebra if ri % 2 == 0 else self.styles.fill_white
            for ci, val in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.font = self.styles.font_td
                cell.fill = fill
                cell.border = self.styles.border_row
                # Right-align numbers
                if isinstance(val, (int, float)):
                    cell.alignment = self.styles.align_right
                    cell.font = self.styles.font_td_num
                else:
                    cell.alignment = self.styles.align_left
            ws.row_dimensions[row].height = ROW_HEIGHTS["td"]
            row += 1

        # Apply column widths
        if col_widths:
            for ci, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

        # spacer
        ws.row_dimensions[row].height = ROW_HEIGHTS["spacer"]
        return row + 1

    def _embed_chart(self, ws, png_bytes: bytes, row: int, col: int = 1) -> int:
        """Embed a PNG chart image into the worksheet. Returns next row."""
        if not png_bytes:
            return row
        try:
            self._chart_counter += 1
            fname = os.path.join(self._tmp_dir, f"chart_{self._chart_counter}.png")
            with open(fname, "wb") as f:
                f.write(png_bytes)
            img = XlImage(fname)
            # Scale to ~750px wide
            img.width = 750
            img.height = int(img.height * (750 / max(img.width, 1)))
            anchor = f"{get_column_letter(col)}{row}"
            ws.add_image(img, anchor)
            # Estimate rows consumed (approx 15px per row)
            rows_consumed = max(1, img.height // 15)
            return row + rows_consumed + 1
        except Exception as e:
            logger.warning("Failed to embed chart: %s", e)
            return row + 1

    def _write_insight(self, ws, key: str, row: int) -> int:
        """Write AI insight block. Returns next row."""
        text = self.insights.get(key, "")
        if not text:
            return row

        title_label = "AI Insight" if self.language == "en" else "AI 洞察"
        row = self._write_section(ws, title_label, row)

        # Write insight text (may be multi-line)
        lines = text.strip().split("\n")
        for line in lines:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
            cell = ws.cell(row=row, column=1, value=line.strip())
            cell.font = self.styles.font_insight_body
            cell.fill = self.styles.fill_insight
            cell.alignment = self.styles.align_wrap
            ws.row_dimensions[row].height = ROW_HEIGHTS["insight"]
            row += 1

        ws.row_dimensions[row].height = ROW_HEIGHTS["spacer"]
        return row + 1

    def _safe_chart(self, chart_fn, *args, **kwargs) -> Optional[bytes]:
        """Call a chart function, return None on failure."""
        try:
            return chart_fn(*args, **kwargs)
        except Exception as e:
            logger.warning("Chart %s failed: %s", chart_fn.__name__, e)
            return None

    # ─── Sheet builders ──────────────────────────────────────────────────

    def _build_sheet_executive(self, ws):
        """Sheet 1 — Executive Summary."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[0])

        # KPI summary table
        row = self._write_section(ws, "KPI Summary" if self.language == "en" else "KPI 概览", row)
        r = self.result
        kpis = r.kpis if isinstance(r.kpis, dict) else {}

        def _kpi_val(key, attr="current_value", default=0):
            m = kpis.get(key)
            if m is None:
                return default
            return getattr(m, attr, default)

        headers = ["KPI", "Value", "Status"]
        data = [
            ["Health Score", format_number(r.health_score), r.health_grade],
            ["Incident SLA Rate", format_pct(_kpi_val("incident_sla_rate")), rating_text(_kpi_val("incident_sla_rate"), self.language)],
            ["MTTR", format_duration(_kpi_val("mttr") * 60, self.language), ""],
            ["Change Success Rate", format_pct(_kpi_val("change_success_rate")), rating_text(_kpi_val("change_success_rate"), self.language)],
            ["Emergency Change Rate", format_pct(_kpi_val("emergency_change_rate")), ""],
            ["Request Fulfillment", format_pct(_kpi_val("request_fulfillment_rate")), rating_text(_kpi_val("request_fulfillment_rate"), self.language)],
            ["CSAT", f"{_kpi_val('csat'):.2f}", ""],
            ["Problem Closure Rate", format_pct(_kpi_val("problem_closure_rate")), rating_text(_kpi_val("problem_closure_rate"), self.language)],
        ]
        row = self._write_table(ws, headers, data, row, [COL_WIDTHS["long_text"], COL_WIDTHS["short_text"], COL_WIDTHS["short_text"]])

        # Charts
        png = self._safe_chart(chart_exec_health_gauge, r.health_score, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_exec_process_radar, r, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        trends_data = {}
        if isinstance(r.trends, dict):
            for k, v in r.trends.items():
                if hasattr(v, "points"):
                    trends_data[k] = [p.value for p in v.points] if v.points else []
                else:
                    trends_data[k] = []
        png = self._safe_chart(chart_exec_sparklines, trends_data, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        # Risk table
        if r.top_risks:
            row = self._write_section(ws, "Top Risks" if self.language == "en" else "主要风险", row)
            risk_headers = ["Priority", "Risk", "Impact", "Process"]
            risk_data = [
                [risk.priority, risk.message, risk.impact, risk.process]
                for risk in r.top_risks[:5]
            ]
            row = self._write_table(ws, risk_headers, risk_data, row,
                                    [COL_WIDTHS["short_text"], COL_WIDTHS["long_text"], COL_WIDTHS["long_text"], COL_WIDTHS["short_text"]])

        row = self._write_insight(ws, "executive_summary", row)

    def _build_sheet_incidents(self, ws):
        """Sheet 2 — Incident Analysis."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[1])

        # Priority breakdown
        priority_rows = self.detail.priority_breakdown()
        if priority_rows:
            row = self._write_section(ws, "Priority Breakdown" if self.language == "en" else "优先级分布", row)
            headers = ["Priority", "Count", "%", "Cum%", "Avg Resolve", "Median", "Min", "Max", "Resp SLA", "Res SLA", "Violations"]
            data = [
                [p.priority, p.count, format_pct(p.pct), format_pct(p.cum_pct),
                 format_duration(p.avg_resolution_min, self.language),
                 format_duration(p.median_resolution_min, self.language),
                 format_duration(p.min_resolution_min, self.language),
                 format_duration(p.max_resolution_min, self.language),
                 format_pct(p.response_sla_rate), format_pct(p.resolution_sla_rate), p.violation_count]
                for p in priority_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # Category breakdown
        category_rows = self.detail.category_breakdown()
        if category_rows:
            row = self._write_section(ws, "Category Breakdown" if self.language == "en" else "分类分布", row)
            headers = ["Category", "Count", "%", "Cum%", "Avg Resolve", "Median", "Std Dev"]
            data = [
                [c.category, c.count, format_pct(c.pct), format_pct(c.cum_pct),
                 format_duration(c.avg_resolution_min, self.language),
                 format_duration(c.median_resolution_min, self.language),
                 format_duration(c.std_resolution_min, self.language)]
                for c in category_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # Charts
        monthly = self.detail.monthly_trends()
        png = self._safe_chart(chart_inc_monthly_trend, monthly, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_inc_priority_pie, priority_rows, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_inc_category_top10, category_rows, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_inc_mttr_boxplot, self.incidents_df, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_inc_p1p2_trend, monthly, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        row = self._write_insight(ws, "incident_detail", row)

    def _build_sheet_sla(self, ws):
        """Sheet 3 — SLA Analysis."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[2])

        # SLA summary table from result.sla_breakdown
        sla_bd = self.result.sla_breakdown or []
        if sla_bd:
            row = self._write_section(ws, "SLA Summary" if self.language == "en" else "SLA 概览", row)
            headers = ["Priority", "Response Rate", "Resolution Rate", "Total", "Violations"]
            data = []
            for s in sla_bd:
                data.append([
                    getattr(s, "priority", ""),
                    format_pct(getattr(s, "response_rate", 0)),
                    format_pct(getattr(s, "resolution_rate", 0)),
                    getattr(s, "total", 0),
                    getattr(s, "violations", 0),
                ])
            row = self._write_table(ws, headers, data, row)

        # Violation list (top 20)
        violations = self.detail.sla_violations()
        if violations:
            row = self._write_section(ws, "SLA Violations (Top 20)" if self.language == "en" else "SLA 违规 (前20)", row)
            headers = ["Order#", "Priority", "Category", "Type", "Overtime", "Resolver", "Status", "Reason"]
            data = [
                [v.order_number, v.priority, v.category, v.violation_type,
                 v.overtime, v.resolver, v.status, v.reason]
                for v in violations[:20]
            ]
            row = self._write_table(ws, headers, data, row)

        # Root cause table
        root_causes = self.detail.violation_root_causes(violations)
        if root_causes:
            row = self._write_section(ws, "Violation Root Causes" if self.language == "en" else "违规根因", row)
            headers = ["Cause", "Count", "%", "Typical Case", "Improvement"]
            data = [
                [rc.cause, rc.count, format_pct(rc.pct), rc.typical_case, rc.improvement]
                for rc in root_causes
            ]
            row = self._write_table(ws, headers, data, row)

        # Charts
        resp_rate = None
        res_rate = None
        priority_rows = self.detail.priority_breakdown()
        if priority_rows:
            total_resp = sum(p.count * p.response_sla_rate for p in priority_rows)
            total_res = sum(p.count * p.resolution_sla_rate for p in priority_rows)
            total_cnt = sum(p.count for p in priority_rows)
            if total_cnt:
                resp_rate = total_resp / total_cnt
                res_rate = total_res / total_cnt

        png = self._safe_chart(chart_sla_gauge_response, resp_rate, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_sla_gauge_resolution, res_rate, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        monthly = self.detail.monthly_trends()
        png = self._safe_chart(chart_sla_monthly_trend, monthly, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_sla_violation_by_priority, priority_rows, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_sla_violation_heatmap, violations, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        row = self._write_insight(ws, "sla_detail", row)

    def _build_sheet_changes(self, ws):
        """Sheet 4 — Change Analysis."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[3])

        # Change type table
        type_rows = self.detail.change_type_breakdown()
        if type_rows:
            row = self._write_section(ws, "Change Type Breakdown" if self.language == "en" else "变更类型分布", row)
            headers = ["Type", "Count", "%", "Success Rate", "Incident Rate", "Avg Duration(h)"]
            data = [
                [t.change_type, t.count, format_pct(t.pct), format_pct(t.success_rate),
                 format_pct(t.incident_rate), f"{t.avg_duration_hours:.1f}"]
                for t in type_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # Change category table
        cat_rows = self.detail.change_category_breakdown()
        if cat_rows:
            row = self._write_section(ws, "Change Category Breakdown" if self.language == "en" else "变更分类分布", row)
            headers = ["Category", "Count", "Success Rate", "Failures", "Incidents", "Risk"]
            data = [
                [c.category, c.count, format_pct(c.success_rate),
                 c.failure_count, c.incident_count, c.risk_level]
                for c in cat_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # Charts
        png = self._safe_chart(chart_chg_type_pie, type_rows, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        monthly = self.detail.monthly_trends()
        png = self._safe_chart(chart_chg_success_trend, monthly, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_chg_category_bar, cat_rows, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_chg_incident_scatter, self.changes_df, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_chg_planning_accuracy, monthly, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        row = self._write_insight(ws, "change_detail", row)

    def _build_sheet_requests(self, ws):
        """Sheet 5 — Request Analysis."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[4])

        # Request type table
        type_rows = self.detail.request_type_breakdown()
        if type_rows:
            row = self._write_section(ws, "Request Type Breakdown" if self.language == "en" else "请求类型分布", row)
            headers = ["Type", "Count", "%", "Completion Rate", "Avg Fulfill(h)", "CSAT"]
            data = [
                [t.request_type, t.count, format_pct(t.pct), format_pct(t.completion_rate),
                 f"{t.avg_fulfillment_hours:.1f}", f"{t.csat:.2f}"]
                for t in type_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # CSAT distribution
        csat_rows = self.detail.csat_distribution()
        if csat_rows:
            row = self._write_section(ws, "CSAT Distribution" if self.language == "en" else "满意度分布", row)
            headers = ["Score", "Label", "Count", "%", "Cum%"]
            data = [
                [c.score, c.label, c.count, format_pct(c.pct), format_pct(c.cum_pct)]
                for c in csat_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # Charts
        png = self._safe_chart(chart_req_type_pie, type_rows, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_req_csat_bar, csat_rows, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        monthly = self.detail.monthly_trends()
        png = self._safe_chart(chart_req_monthly_trend, monthly, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_req_fulfillment_boxplot, self.requests_df, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_req_dept_heatmap, self.requests_df, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        row = self._write_insight(ws, "request_detail", row)

    def _build_sheet_problems(self, ws):
        """Sheet 6 — Problem Analysis."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[5])

        # Problem status table
        status_rows = self.detail.problem_status_breakdown()
        if status_rows:
            row = self._write_section(ws, "Problem Status" if self.language == "en" else "问题状态", row)
            headers = ["Status", "Count", "%", "Avg Age(days)", "Suggestion"]
            data = [
                [s.status, s.count, format_pct(s.pct), f"{s.avg_age_days:.0f}", s.suggestion]
                for s in status_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # Root cause category table
        rc_rows = self.detail.root_cause_category_breakdown()
        if rc_rows:
            row = self._write_section(ws, "Root Cause Categories" if self.language == "en" else "根因分类", row)
            headers = ["Category", "Count", "%", "Related Incidents", "Fix Rate", "Typical Problem"]
            data = [
                [rc.category, rc.count, format_pct(rc.pct), rc.related_incidents,
                 format_pct(rc.permanent_fix_rate), rc.typical_problem]
                for rc in rc_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # Charts
        png = self._safe_chart(chart_prb_status_funnel, status_rows, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_prb_rootcause_pie, rc_rows, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        monthly = self.detail.monthly_trends()
        png = self._safe_chart(chart_prb_monthly_bar, monthly, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_prb_impact_bubble, self.problems_df, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_prb_rca_trend, monthly, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        row = self._write_insight(ws, "problem_detail", row)

    def _build_sheet_cross_process(self, ws):
        """Sheet 7 — Cross-Process Analysis."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[6])

        # Change -> Incident links
        chg_links = self.detail.change_incident_links()
        if chg_links:
            row = self._write_section(ws, "Change → Incident Links" if self.language == "en" else "变更→事件关联", row)
            headers = ["Source ID", "Type", "Target Count", "Target IDs", "Impact"]
            data = [
                [l.source_id, l.source_type, l.target_count, l.target_ids, l.impact]
                for l in chg_links[:20]
            ]
            row = self._write_table(ws, headers, data, row)

        # Problem -> Incident links
        prb_links = self.detail.problem_incident_links()
        if prb_links:
            row = self._write_section(ws, "Problem → Incident Links" if self.language == "en" else "问题→事件关联", row)
            headers = ["Source ID", "Type", "Target Count", "Target IDs", "Impact"]
            data = [
                [l.source_id, l.source_type, l.target_count, l.target_ids, l.impact]
                for l in prb_links[:20]
            ]
            row = self._write_table(ws, headers, data, row)

        # Charts
        png = self._safe_chart(chart_cross_sankey, chg_links, prb_links, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_cross_radar, [], self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_cross_timeline, chg_links, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_cross_heatmap, None, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        row = self._write_insight(ws, "cross_process", row)

    def _build_sheet_personnel(self, ws):
        """Sheet 8 — Personnel & Efficiency."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[7])

        # Personnel table
        personnel = self.detail.personnel_breakdown()
        if personnel:
            row = self._write_section(ws, "Personnel Performance" if self.language == "en" else "人员绩效", row)
            headers = ["Name", "Count", "%", "Avg Resolve", "Completion", "Resp SLA", "Res SLA", "Rating", "Specialty"]
            data = [
                [p.name, p.count, format_pct(p.pct),
                 format_duration(p.avg_resolution_min, self.language),
                 format_pct(p.completion_rate),
                 format_pct(p.response_sla_rate), format_pct(p.resolution_sla_rate),
                 p.rating, p.specialty]
                for p in personnel
            ]
            row = self._write_table(ws, headers, data, row)

        # Workload distribution
        workload = self.detail.workload_distribution(personnel)
        if workload:
            row = self._write_section(ws, "Workload Distribution" if self.language == "en" else "工作量分布", row)
            headers = ["Level", "Count", "%", "Avg Events", "Suggestion"]
            data = [
                [w.level, w.count, format_pct(w.pct), f"{w.avg_events:.1f}", w.suggestion]
                for w in workload
            ]
            row = self._write_table(ws, headers, data, row)

        # Skill coverage
        skill = self.detail.skill_coverage()
        if skill:
            row = self._write_section(ws, "Skill Coverage" if self.language == "en" else "技能覆盖", row)
            headers = ["Category", "Handlers", "Primary", "Coverage", "Risk"]
            data = [
                [s.category, s.handler_count, s.primary_handler,
                 format_pct(s.coverage_rate), s.risk_level]
                for s in skill
            ]
            row = self._write_table(ws, headers, data, row)

        # Charts
        png = self._safe_chart(chart_pers_workload_bar, personnel, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_pers_load_boxplot, personnel, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_pers_skill_heatmap, self.incidents_df, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_pers_efficiency_scatter, personnel, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_pers_top10_radar, personnel, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        row = self._write_insight(ws, "personnel", row)

    def _build_sheet_time(self, ws):
        """Sheet 9 — Time Analysis."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[8])

        # Monthly trends table
        monthly = self.detail.monthly_trends()
        if monthly:
            row = self._write_section(ws, "Monthly Trends" if self.language == "en" else "月度趋势", row)
            headers = ["Period", "Incidents", "Changes", "Requests", "Problems",
                        "Completion Rate", "Avg Resolve", "High Pri%", "MoM", "Assessment"]
            data = [
                [m.period, m.incident_count, m.change_count, m.request_count, m.problem_count,
                 format_pct(m.completion_rate),
                 format_duration(m.avg_resolution_min, self.language),
                 format_pct(m.high_priority_pct), m.mom_change, m.assessment]
                for m in monthly
            ]
            row = self._write_table(ws, headers, data, row)

        # Day-of-week table
        dow = self.detail.day_of_week_analysis()
        if dow:
            row = self._write_section(ws, "Day of Week" if self.language == "en" else "星期分布", row)
            headers = ["Day", "Count", "%", "Avg Resolve", "High Pri", "Assessment"]
            data = [
                [d.day, d.count, format_pct(d.pct),
                 format_duration(d.avg_resolution_min, self.language),
                 d.high_priority_count, d.assessment]
                for d in dow
            ]
            row = self._write_table(ws, headers, data, row)

        # Hour-of-day table
        hod = self.detail.hour_of_day_analysis()
        if hod:
            row = self._write_section(ws, "Hour of Day" if self.language == "en" else "时段分布", row)
            headers = ["Period", "Hours", "Count", "%", "Avg Resolve", "Suggestion"]
            data = [
                [h.period, h.hour_range, h.count, format_pct(h.pct),
                 format_duration(h.avg_resolution_min, self.language), h.suggestion]
                for h in hod
            ]
            row = self._write_table(ws, headers, data, row)

        # Charts
        png = self._safe_chart(chart_time_four_process_trend, monthly, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_time_dow_bar, dow, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_time_hour_heatmap, self.incidents_df, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_time_quarterly, monthly, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_time_forecast, monthly, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        row = self._write_insight(ws, "time_analysis", row)

    def _build_sheet_actions(self, ws):
        """Sheet 10 — Action Plan."""
        names = SHEET_NAMES[self.language]
        row = self._write_title(ws, names[9])

        # Build ActionPlanRow list from result.actions
        actions = self.result.actions or []
        action_plan_rows: List[ActionPlanRow] = []
        for i, a in enumerate(actions, 1):
            action_plan_rows.append(ActionPlanRow(
                seq=i,
                priority=getattr(a, "priority", "Medium"),
                action=getattr(a, "action", ""),
                source_sheet="",
                source_process=getattr(a, "process", "General"),
                responsible="",
                expected_effect=getattr(a, "expected_impact", ""),
            ))

        if action_plan_rows:
            row = self._write_section(ws, "Action Items" if self.language == "en" else "行动项", row)
            headers = ["#", "Priority", "Action", "Process", "Expected Effect"]
            data = [
                [a.seq, a.priority, a.action, a.source_process, a.expected_effect]
                for a in action_plan_rows
            ]
            row = self._write_table(ws, headers, data, row)

        # Risk summary table
        risks = self.result.risks or []
        if risks:
            row = self._write_section(ws, "Risk Summary" if self.language == "en" else "风险汇总", row)
            headers = ["ID", "Priority", "Message", "Impact", "Process"]
            data = [
                [r.id, r.priority, r.message, r.impact, r.process]
                for r in risks
            ]
            row = self._write_table(ws, headers, data, row)

        # Charts
        png = self._safe_chart(chart_action_priority_pie, action_plan_rows, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        png = self._safe_chart(chart_action_process_bar, action_plan_rows, self.language)
        if png:
            row = self._embed_chart(ws, png, row)

        row = self._write_insight(ws, "action_plan", row)

    # ─── Build & Save ────────────────────────────────────────────────────

    def build(self) -> openpyxl.Workbook:
        """Create the workbook with all 10 sheets."""
        self.wb = openpyxl.Workbook()
        names = SHEET_NAMES[self.language]

        builders = [
            self._build_sheet_executive,
            self._build_sheet_incidents,
            self._build_sheet_sla,
            self._build_sheet_changes,
            self._build_sheet_requests,
            self._build_sheet_problems,
            self._build_sheet_cross_process,
            self._build_sheet_personnel,
            self._build_sheet_time,
            self._build_sheet_actions,
        ]

        for i, (name, builder) in enumerate(zip(names, builders)):
            if i == 0:
                ws = self.wb.active
                ws.title = name
            else:
                ws = self.wb.create_sheet(title=name)

            try:
                builder(ws)
            except Exception as e:
                logger.error("Failed to build sheet '%s': %s", name, e)
                ws.cell(row=1, column=1, value=f"Error building sheet: {e}")

        # Apply print settings
        for ws in self.wb.worksheets:
            ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
                fitToPage=True
            )
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0

        return self.wb

    def save(self, filename: str = None) -> Path:
        """Build workbook, save to OUTPUT_DIR, clean up temp files, return path."""
        if self.wb is None:
            self.build()

        if filename is None:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Comprehensive_Quality_Report_{ts}.xlsx"

        output_path = OUTPUT_DIR / filename
        self.wb.save(str(output_path))

        # Clean up temp chart images
        try:
            shutil.rmtree(self._tmp_dir, ignore_errors=True)
        except Exception:
            pass

        logger.info("Report saved to %s", output_path)
        return output_path
